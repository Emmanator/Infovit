## Program som summerer kolonnene i en tallmatrise
import openpyxl
from openpyxl.utils import get_column_letter
##from openpyxl.cell.cell import get_column_letter ## alternativ

tall=openpyxl.load_workbook('tallmatrise.xlsx')


matrise=tall.active

antallRader=matrise.max_row
antallKolonner=matrise.max_column

sumRad=str(antallRader+2)

for k in range(1, antallKolonner+1):
    bokstav=get_column_letter(k)
    funk='=SUM('+bokstav+'1'+':'+bokstav+str(antallRader)+')'
    matrise[bokstav+sumRad]=funk

tall.save('tallmatriseSum.xlsx')






    
    

    
    





